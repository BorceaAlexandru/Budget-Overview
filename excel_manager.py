import openpyxl
from openpyxl import Workbook
import os
import uuid
from datetime import date
from utils import get_week_label, adjust_day_for_month, get_filename

class ExcelManager:
    def __init__(self, year, month):
        self.year = year
        self.month = month
        self.filename = get_filename(year, month)

    def _ensure_file_exists(self):
        if not os.path.exists(self.filename):
            wb=Workbook()
            if "Sheet" not in wb.sheetnames:
                del wb["Sheet"]

            ws_tr=wb.create_sheet("Tranzactii")
            ws_tr.append(["ID", "Data", "Saptamana", "Categorie", "Suma", "Descriere", "Tip", "Sursa"])
            wb.create_sheet("SumarSaptamani")
            wb.create_sheet("CategorieXSaptamana")
            wb.save(self.filename)

    def add_transaction(self, day, categorie, suma, descriere, tip, sursa):
        self._ensure_file_exists()
        wb=openpyxl.load_workbook(self.filename)
        ws=wb["Tranzactii"]

        valid_day=adjust_day_for_month(self.year, self.month, day)
        tr_date=date(self.year, self.month, valid_day)
        sapt=get_week_label(valid_day)
        tr_id=str(uuid.uuid4())[:8]

        ws.append([tr_id, tr_date.strftime("%Y-%m-%d"), sapt, categorie, suma, descriere, tip, sursa])
        wb.save(self.filename)
        self.recalculate_summaries()

    def process_fixed_expenses(self, templates):
        self._ensure_file_exists()
        wb=openpyxl.load_workbook(self.filename)
        ws=wb["Tranzactii"]

        existing_fixed=set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            #row: 0=ID, 1=Date, 2=Sapt, 3=Cat, 4=Suma, 5=Desc, 6=Tip, 7=Sursa
            if row[6]=="FIXA" and row[7]=="TEMPLATE":
                existing_fixed.add(row[5])

        added_count=0
        for t in templates:
            if t['nume'] not in existing_fixed:
                valid_day=adjust_day_for_month(self.year, self.month, t['zi'])
                tr_date=date(self.year, self.month, valid_day)
                sapt=get_week_label(valid_day)
                tr_id=str(uuid.uuid4())[:8]

                ws.append([
                    tr_id,
                    tr_date.strftime("%Y-%m-%d"),
                    sapt,
                    t['categorie'],
                    t['suma'],
                    t['nume'],
                    "FIXA",
                    "TEMPLATE"
                ])
                added_count+=1

        wb.save(self.filename)
        if added_count > 0:
            self.recalculate_summaries()
        return added_count

    def recalculate_summaries(self):
        wb=openpyxl.load_workbook(self.filename)
        ws_tr=wb["Tranzactii"]

        weeks=["S1", "S2", "S3", "S4", "S5"]
        matrix={}
        weekly_totals ={w: 0.0 for w in weeks}

        for row in ws_tr.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue

            sapt=row[2]
            cat=row[3]
            try:
                suma=float(row[4])
            except:
                suma=0.0

            if cat not in matrix:
                matrix[cat]={w: 0.0 for w in weeks}

            if sapt in weeks:
                matrix[cat][sapt]+=suma
                weekly_totals[sapt]+=suma

        ws_sum=wb["SumarSaptamani"]
        ws_sum.delete_rows(1, ws_sum.max_row + 1)

        ws_sum.append(["Saptamana", "Total Cheltuit"])
        total_lunar=0
        for w in weeks:
            ws_sum.append([w, weekly_totals[w]])
            total_lunar+=weekly_totals[w]
        ws_sum.append(["TOTAL LUNAR", total_lunar])

        ws_cat=wb["CategorieXSaptamana"]
        ws_cat.delete_rows(1, ws_cat.max_row + 1)

        header=["Categorie"]+weeks+["Total Categorie"]
        ws_cat.append(header)

        for cat in sorted(matrix.keys()):
            row_data=[cat]
            cat_total=0
            for w in weeks:
                val=matrix[cat][w]
                row_data.append(val)
                cat_total+=val
            row_data.append(cat_total)
            ws_cat.append(row_data)

        wb.save(self.filename)